import openpyxl


wb = openpyxl.Workbook()
ws = wb.create_sheet('second sheet')
ws2 = wb.create_sheet('first sheet', 0)
ws.title = 'the sheet'
ws.sheet_properties.tabColor = "1072BA"
ws3 = wb['the sheet']
# ws and ws3 are the same
if ws == ws3:
    print("\n they are the same \n")
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)
target = wb.copy_worksheet(ws)
print(wb.sheetnames)
c = ws['A4']
ws['A4'] = '4'
c2 = ws.cell(row=4, column=2, value='20')
# the following code will show an example of creating a memory to the cells by scrolling into them
for x in range(1, 101):
    for y in range(1, 101):
        ws.cell(row=x, column=y)
cell_range = ws['A1':'C2']
# another way to range cells
row_range1 = ws[5:10]
col_range1 = ws['A':'D']
# another way of iterating rows
for row in ws.iter_rows(min_row=1,  max_col=3, max_row=2):
    print(row)
# other way but for iterating between columns
for column in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    print(column)
print(wb.active)
ws = wb.active
ws['C9'] = 'none sense'
print(tuple(ws.rows))
print(tuple(ws.columns))
print('\n'*10)
for row in ws.values:
    for values in row:
        print(values)
print(ws.values)
wb.save("myfirst workbook.xlsx")
